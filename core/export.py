"""Excel export for all measurement sequences (openpyxl).

NOTE: openpyxl writes .xlsx natively.  True .xlsm (VBA macros) requires an
existing macro-enabled template opened with keep_vba=True.  For new files we
use .xlsx — Excel opens these identically for pure data + charts.
"""
from __future__ import annotations

import os
import statistics
from datetime import datetime
from typing import Any, Dict, List

import openpyxl
from openpyxl.chart import LineChart, Reference, ScatterChart
from openpyxl.chart.series_factory import SeriesFactory as Series
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from .equipment.base import MeasureResult
from .gamut_utils import calc_gamut_stats, DCI_P3_UV, BT2020_UV

# ── 브랜드 색상 ───────────────────────────────────────────────────────────────
_BRAND_HEX = {"samsung": "0070C0", "lg": "FF0000", "sony": "00B050", "philips": "FF8800"}


def _brand_hex(brand: str) -> str:
    return _BRAND_HEX.get(brand.lower().strip(), "FF8800")


# ── 스타일 상수 ──────────────────────────────────────────────────────────────
_BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
_GRAY_FILL   = PatternFill("solid", fgColor="D6DCE4")
_GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
_YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
_RED_FILL    = PatternFill("solid", fgColor="FCE4D6")

_WHITE_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
_BOLD_FONT   = Font(bold=True, name="Calibri", size=10)
_BASE_FONT   = Font(name="Calibri", size=10)

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")

# 명암비 값 → 배경색 (높을수록 짙은 녹색)
def _cr_fill(cr: float) -> PatternFill:
    if   cr >= 8000: return PatternFill("solid", fgColor="1E6B44")
    elif cr >= 4000: return PatternFill("solid", fgColor="375623")
    elif cr >= 2000: return PatternFill("solid", fgColor="70AD47")
    elif cr >= 1000: return PatternFill("solid", fgColor="C6EFCE")
    elif cr >=  500: return PatternFill("solid", fgColor="FFEB9C")
    else:            return PatternFill("solid", fgColor="FFC7CE")


def _cr_font(cr: float) -> Font:
    if   cr >= 2000: return Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    elif cr >= 1000: return Font(bold=True, color="375623", name="Calibri", size=10)
    elif cr >=  500: return Font(bold=True, color="9C5700", name="Calibri", size=10)
    else:            return Font(bold=True, color="9C0006", name="Calibri", size=10)


# 숫자 형식
_FMT_LV    = '0.000'
_FMT_CHROM = '0.0000'
_FMT_XYZ   = '0.000'
_FMT_INT   = '0'
_FMT_RATIO = '0.0'

# 측정값 컬럼 정의: (헤더, 속성경로, 숫자포맷, 열너비)
_MEAS_COLS: list[tuple[str, str, str, int]] = [
    ("Time (s)",        "timestamp_ms",           _FMT_INT,   10),
    ("Lv (cd/m²)",      "Lv",                     _FMT_LV,    12),
    ("x",               "x",                      _FMT_CHROM, 10),
    ("y",               "y",                      _FMT_CHROM, 10),
    ("u'",              "u_prime",                _FMT_CHROM, 10),
    ("v'",              "v_prime",                _FMT_CHROM, 10),
    ("X",               "X",                      _FMT_XYZ,   10),
    ("Y",               "Y",                      _FMT_XYZ,   10),
    ("Z",               "Z",                      _FMT_XYZ,   10),
    ("CCT (K)",         "cct",                    _FMT_INT,   10),
    ("Duv",             "duv",                    "0.00000",  10),
    ("Pattern",         "pattern_info.type",      "@",         12),
    ("APL (%)",         "pattern_info.apl_pct",   _FMT_RATIO, 10),
    ("W (%)",           "pattern_info.width_pct", _FMT_RATIO, 10),
    ("H (%)",           "pattern_info.height_pct",_FMT_RATIO, 10),
    ("Color",           "pattern_info.color",     "@",         12),
    ("SDR/HDR",         "pattern_info.is_hdr",    "@",         10),
]


# ── 헬퍼 ─────────────────────────────────────────────────────────────────────

def _get_attr(obj: object, path: str) -> object:
    for part in path.split("."):
        obj = getattr(obj, part)
    if isinstance(obj, bool):
        return "HDR" if obj else "SDR"
    return obj


def _write_header_row(ws, row: int, labels: list[str],
                      fill=_BLUE_FILL, font=_WHITE_FONT) -> None:
    for ci, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.fill = fill
        c.font = font
        c.alignment = _CENTER
        c.border = _BORDER


def _write_meas_row(ws, row: int, r: MeasureResult,
                    col_offset: int = 0, seq: int | None = None) -> None:
    for ci, (_, path, fmt, _) in enumerate(_MEAS_COLS, 1 + col_offset):
        val = seq if (seq is not None and path == "timestamp_ms") else _get_attr(r, path)
        c = ws.cell(row=row, column=ci, value=val)
        c.number_format = fmt
        c.alignment = _CENTER
        c.border = _BORDER
        c.font = _BASE_FONT


def _set_col_widths(ws, widths: list[int], col_offset: int = 0) -> None:
    for ci, w in enumerate(widths, 1 + col_offset):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _freeze(ws, cell: str = "B2") -> None:
    ws.freeze_panes = cell


def _file_prefix(brand: str, model: str) -> str:
    safe = lambda s: "".join(c for c in s if c.isalnum() or c in "._- ").strip()
    return f"{safe(brand)}_{safe(model)}"


def _default_path(brand: str, model: str, suffix: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{_file_prefix(brand, model)}_{suffix}_{ts}.xlsx"
    return os.path.join(os.getcwd(), fname)


def _add_info_sheet(
    wb,
    brand: str,
    model: str,
    sequence: str,
    serial_number: str = "",
    sw_version: str = "",
    sw_codename: str = "",
) -> None:
    # openpyxl.Workbook() always creates a default empty "Sheet" — remove it
    for default_name in ("Sheet", "Sheet1"):
        if default_name in wb.sheetnames:
            wb.remove(wb[default_name])
    ws = wb.create_sheet("Info", 0)
    rows = [
        ("Brand",     brand),
        ("Model",     model),
        ("Sequence",  sequence),
        ("Export",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Software",  "LED Measure v0.1"),
    ]
    if serial_number:
        rows.append(("Serial Number", serial_number))
    if sw_version:
        rows.append(("SW Version", sw_version))
    if sw_codename:
        rows.append(("SW Codename", sw_codename))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    for ri, (k, v) in enumerate(rows, 1):
        ca = ws.cell(row=ri, column=1, value=k)
        cb = ws.cell(row=ri, column=2, value=v)
        ca.font = _BOLD_FONT
        cb.font = _BASE_FONT
        ca.border = _BORDER
        cb.border = _BORDER


# ── ExcelExporter ─────────────────────────────────────────────────────────────

class ExcelExporter:
    """Creates formatted Excel workbooks from MeasureResult data."""

    # ── 1. Luminance Swing ────────────────────────────────────────────────────

    def export_lum_swing(
        self,
        results_by_case: Dict[str, List[MeasureResult]],
        brand: str,
        model: str,
        file_path: str | None = None,
    ) -> str:
        """단일 시트 피벗: 행=측정#, 열=모드(SDR_Vivid 등), 값=Lv."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _add_info_sheet(wb, brand, model, "Luminance Swing")

        ws = wb.create_sheet("LumSwing")
        keys = sorted(results_by_case.keys())
        max_n = max((len(v) for v in results_by_case.values()), default=0)

        # 헤더: # | SDR_Vivid | SDR_Standard | ...
        _write_header_row(ws, 1, ["#"] + keys)
        ws.column_dimensions["A"].width = 6
        for ci in range(2, len(keys) + 2):
            ws.column_dimensions[get_column_letter(ci)].width = 14
        _freeze(ws, "B2")

        # 데이터 행
        for i in range(max_n):
            ri = i + 2
            idx_cell = ws.cell(row=ri, column=1, value=i + 1)
            idx_cell.alignment = _CENTER
            idx_cell.border = _BORDER
            idx_cell.font = _BASE_FONT
            for ci, key in enumerate(keys, 2):
                row_list = results_by_case.get(key, [])
                val = row_list[i].Lv if i < len(row_list) else None
                c = ws.cell(row=ri, column=ci, value=val)
                c.number_format = _FMT_LV
                c.alignment = _CENTER
                c.border = _BORDER
                c.font = _BASE_FONT

        # 통계 블록
        if max_n > 0:
            stat_row = max_n + 3
            ws.cell(row=stat_row, column=1, value="Statistics").font = _BOLD_FONT
            stat_row += 1
            for label in ["Count", "Mean Lv", "Max Lv", "Min Lv", "Std Dev"]:
                ws.cell(row=stat_row, column=1, value=label).font = _BOLD_FONT
                for ci, key in enumerate(keys, 2):
                    lvs = [r.Lv for r in results_by_case.get(key, [])]
                    if not lvs:
                        continue
                    if   label == "Count":   val = len(lvs)
                    elif label == "Mean Lv": val = round(statistics.mean(lvs), 3)
                    elif label == "Max Lv":  val = round(max(lvs), 3)
                    elif label == "Min Lv":  val = round(min(lvs), 3)
                    else:                    val = round(statistics.stdev(lvs), 4) if len(lvs) > 1 else 0
                    c = ws.cell(row=stat_row, column=ci, value=val)
                    c.number_format = _FMT_LV
                    c.alignment = _CENTER
                    c.border = _BORDER
                stat_row += 1

        if max_n >= 2:
            self._add_swing_pivot_chart(ws, keys, max_n)

        path = file_path or _default_path(brand, model, "LumSwing")
        wb.save(path)
        return path

    def _add_swing_pivot_chart(self, ws, keys: list, data_rows: int) -> None:
        if data_rows < 2:
            return
        chart = LineChart()
        chart.title = "Luminance Swing"
        chart.style = 10
        chart.y_axis.title = "Lv (cd/m²)"
        chart.x_axis.title = "측정 #"
        chart.width  = 26
        chart.height = 14

        _COLORS = ["1F4E79", "C55A11", "375623", "7030A0", "FF0000", "00B0F0"]
        for i, key in enumerate(keys):
            ref = Reference(ws, min_col=i + 2, min_row=1, max_row=data_rows + 1)
            chart.add_data(ref, titles_from_data=True)
            chart.series[i].graphicalProperties.line.solidFill = _COLORS[i % len(_COLORS)]
            chart.series[i].graphicalProperties.line.width = 15000

        cat_ref = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)
        chart.set_categories(cat_ref)
        ws.add_chart(chart, f"A{data_rows + 6}")

    # ── 2. Luminance Loading ──────────────────────────────────────────────────

    def export_lum_loading(
        self,
        results_by_case: Dict[str, Dict[int, List[MeasureResult]]],
        brand: str,
        model: str,
        use_avg: bool = True,   # kept for API compatibility; Summary shows both Avg & Max
        file_path: str | None = None,
        brand_name: str = "",
    ) -> str:
        """Summary 시트(APL × 케이스) + 케이스별 Raw 시트."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _add_info_sheet(wb, brand, model, "Luminance Loading")

        cases   = sorted(results_by_case.keys())
        all_apls: list[int] = sorted(
            {apl for cd in results_by_case.values() for apl in cd}
        )

        # ── Summary 시트 ──────────────────────────────────────────────────────
        ws_sum = wb.create_sheet("Summary")

        # 헤더: APL | Case A Avg | Case A Max | Case A Min | Case B ...
        hdr = ["APL (%)"]
        for case in cases:
            hdr += [f"Case {case} Avg Lv", f"Case {case} Max Lv",
                    f"Case {case} Min Lv", f"Case {case} StdDev"]
        _write_header_row(ws_sum, 1, hdr)
        ws_sum.column_dimensions["A"].width = 10
        for ci in range(2, len(hdr) + 1):
            ws_sum.column_dimensions[get_column_letter(ci)].width = 14

        for ri, apl in enumerate(all_apls, 2):
            ws_sum.cell(row=ri, column=1, value=apl).alignment = _CENTER
            ci = 2
            for case in cases:
                step_results = results_by_case.get(case, {}).get(apl, [])
                lvs = [r.Lv for r in step_results]
                avg = round(statistics.mean(lvs), 3)   if lvs else None
                mx  = round(max(lvs), 3)               if lvs else None
                mn  = round(min(lvs), 3)               if lvs else None
                std = round(statistics.stdev(lvs), 4)  if len(lvs) > 1 else (0 if lvs else None)
                for val, fmt in [(avg, _FMT_LV), (mx, _FMT_LV), (mn, _FMT_LV), (std, _FMT_LV)]:
                    c = ws_sum.cell(row=ri, column=ci, value=val)
                    c.number_format = fmt
                    c.alignment = _CENTER
                    c.border = _BORDER
                    ci += 1

        _freeze(ws_sum, "B2")

        # Summary 라인 차트 (Avg Lv per case)
        self._add_apl_chart(ws_sum, all_apls, cases, len(all_apls), brand=brand_name)

        # ── Raw 시트 ─────────────────────────────────────────────────────────
        meas_headers = [h for h, *_ in _MEAS_COLS]
        meas_widths  = [w for *_, w in _MEAS_COLS]

        for case, case_data in sorted(results_by_case.items()):
            ws_raw = wb.create_sheet(title=f"Raw_{case}")
            _write_header_row(ws_raw, 1, ["APL (%)"] + meas_headers,
                              fill=PatternFill("solid", fgColor="375623"),
                              font=_WHITE_FONT)
            ws_raw.column_dimensions["A"].width = 10
            _set_col_widths(ws_raw, meas_widths, col_offset=1)
            _freeze(ws_raw, "B2")

            ri = 2
            for apl in sorted(case_data.keys()):
                fill = _GRAY_FILL if apl % 20 < 10 else None
                for r in case_data[apl]:
                    c = ws_raw.cell(row=ri, column=1, value=apl)
                    c.alignment = _CENTER
                    c.border = _BORDER
                    if fill:
                        c.fill = fill
                    _write_meas_row(ws_raw, ri, r, col_offset=1, seq=ri - 1)
                    if fill:
                        for ci2 in range(2, len(_MEAS_COLS) + 2):
                            ws_raw.cell(row=ri, column=ci2).fill = fill
                    ri += 1

        path = file_path or _default_path(brand, model, "LumLoading")
        wb.save(path)
        return path

    def _add_apl_chart(self, ws, _all_apls: list, cases: list, n_apl: int,
                       brand: str = "", cols_per_case: int = 4) -> None:
        chart = LineChart()
        chart.title = "APL vs Lv (Average)"
        chart.style = 10
        chart.y_axis.title = "Lv (cd/m²)"
        chart.x_axis.title = "APL (%)"
        chart.width  = 26
        chart.height = 14

        fallback_colors = ["1F4E79", "C55A11", "375623", "7030A0", "FF8800", "00B0F0"]
        for i, case in enumerate(cases):
            col = 2 + i * cols_per_case   # Avg 컬럼
            ref = Reference(ws, min_col=col, min_row=1, max_row=n_apl + 1)
            chart.add_data(ref, titles_from_data=True)
            color = _brand_hex(brand) if i == 0 and brand else (
                fallback_colors[i % len(fallback_colors)]
            )
            chart.series[i].graphicalProperties.line.solidFill = color
            chart.series[i].graphicalProperties.line.width = 18000

        apl_ref = Reference(ws, min_col=1, min_row=2, max_row=n_apl + 1)
        chart.set_categories(apl_ref)
        ws.add_chart(chart, f"A{n_apl + 4}")

    def _add_gamut_chart(self, ws, results: dict, anchor: str) -> None:
        """u'v' scatter chart — 측정 삼각형 + DCI-P3/BT.2020 기준선."""
        _C = 22   # 헬퍼 데이터를 기록할 시작 열 (V열)

        # 닫힌 삼각형 포인트 목록
        dci_pts  = list(DCI_P3_UV) + [DCI_P3_UV[0]]
        bt_pts   = list(BT2020_UV) + [BT2020_UV[0]]
        meas_raw = [results.get(c) for c in ("red", "green", "blue")]
        meas_uvs = [(r.u_prime, r.v_prime) for r in meas_raw if r]
        meas_pts = (meas_uvs + [meas_uvs[0]]) if len(meas_uvs) == 3 else meas_uvs

        # 헬퍼 데이터 기록 (u, v 쌍 × 3 그룹)
        for grp_i, pts in enumerate([dci_pts, bt_pts, meas_pts]):
            bc = _C + grp_i * 2
            for row_i, (u, v) in enumerate(pts, 1):
                ws.cell(row=row_i, column=bc,     value=round(u, 6))
                ws.cell(row=row_i, column=bc + 1, value=round(v, 6))

        n_dci, n_bt, n_meas = len(dci_pts), len(bt_pts), len(meas_pts)

        chart = ScatterChart()
        chart.scatterStyle = "lineMarker"
        chart.title  = "u'v' Color Gamut"
        chart.style  = 2
        chart.x_axis.title = "u'"
        chart.y_axis.title = "v'"
        chart.x_axis.scaling.min = 0.0
        chart.x_axis.scaling.max = 0.65
        chart.y_axis.scaling.min = 0.0
        chart.y_axis.scaling.max = 0.65
        chart.width  = 18
        chart.height = 18

        def _ser(x_col, y_col, n_pts, title, hex_color, marker="none", m_size=5):
            s = Series(
                values  = Reference(ws, min_col=y_col, min_row=1, max_row=n_pts),
                xvalues = Reference(ws, min_col=x_col, min_row=1, max_row=n_pts),
                title   = title,
            )
            s.smooth = False  # 직선 연결 강제 (기본값이 True 이면 Bezier 곡선으로 그려짐)
            s.graphicalProperties.line.solidFill = hex_color
            s.graphicalProperties.line.width = 15000
            s.marker.symbol = marker
            if marker != "none":
                s.marker.size = m_size
                s.marker.graphicalProperties.solidFill       = hex_color
                s.marker.graphicalProperties.line.solidFill = hex_color
            return s

        chart.series.append(_ser(_C,   _C+1, n_dci, "DCI-P3",  "4472C4"))
        chart.series.append(_ser(_C+2, _C+3, n_bt,  "BT.2020", "A6A6A6"))
        if n_meas >= 3:
            chart.series.append(
                _ser(_C+4, _C+5, n_meas, "Measured", "FFC000", marker="circle", m_size=7)
            )
        ws.add_chart(chart, anchor)

    # ── 3. Gamut ──────────────────────────────────────────────────────────────

    def export_gamut(
        self,
        results: Dict[str, MeasureResult],
        brand: str,
        model: str,
        file_path: str | None = None,
        gamut_stats: Dict[str, float] | None = None,
    ) -> str:
        """컬러별 행 + 배경색, 전체 색채측정값."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _add_info_sheet(wb, brand, model, "Gamut")

        ws = wb.create_sheet("Gamut")

        color_order = ["red", "green", "blue", "white", "black"]
        color_fills = {
            "red":   PatternFill("solid", fgColor="FCE4D6"),
            "green": PatternFill("solid", fgColor="E2EFDA"),
            "blue":  PatternFill("solid", fgColor="DAE8FC"),
            "white": PatternFill("solid", fgColor="F2F2F2"),
            "black": PatternFill("solid", fgColor="D9D9D9"),
        }
        color_label = {
            "red": "Red", "green": "Green", "blue": "Blue",
            "white": "White", "black": "Black",
        }

        headers = ["Color"] + [h for h, *_ in _MEAS_COLS]
        widths  = [8] + [w for *_, w in _MEAS_COLS]
        _write_header_row(ws, 1, headers)
        _set_col_widths(ws, widths)
        ws.column_dimensions["A"].width = 28   # 통계 레이블 여유 확보
        _freeze(ws, "B2")

        for ri, color in enumerate(color_order, 2):
            r = results.get(color)
            fill = color_fills.get(color)
            label_cell = ws.cell(row=ri, column=1, value=color_label.get(color, color))
            label_cell.font = _BOLD_FONT
            label_cell.alignment = _CENTER
            label_cell.border = _BORDER
            if fill:
                label_cell.fill = fill

            if r:
                _write_meas_row(ws, ri, r, col_offset=1)
                if fill:
                    for ci in range(2, len(_MEAS_COLS) + 2):
                        ws.cell(row=ri, column=ci).fill = fill

        # White/Black 명암비 계산
        w_r = results.get("white")
        bk_r = results.get("black")
        stat_row = len(color_order) + 3
        if w_r and bk_r and bk_r.Lv > 0:
            cr = round(w_r.Lv / bk_r.Lv, 1)
            ws.cell(row=stat_row, column=1, value="Contrast Ratio").font = _BOLD_FONT
            c = ws.cell(row=stat_row, column=2, value=cr)
            c.number_format = _FMT_RATIO
            c.font = Font(bold=True, color="C55A11", size=11)
            stat_row += 1

        # DCI-P3 / BT.2020 gamut 통계
        if gamut_stats:
            _stat_fill = PatternFill("solid", fgColor="EBF3FB")
            # (label, key, number_format)
            stat_items = [
                ("DCI-P3 Overlap (%)",       "dci_overlap",       "0.00"),
                ("DCI-P3 Intersection Area", "dci_inter_area",    "0.000000"),
                ("DCI-P3 Area",              "dci_area",          "0.000000"),
                ("DCI-P3 Area Ratio (%)",    "dci_area_ratio",    "0.00"),
                ("BT.2020 Overlap (%)",      "bt2020_overlap",    "0.00"),
                ("BT.2020 Intersection Area","bt2020_inter_area", "0.000000"),
                ("BT.2020 Area",             "bt2020_area",       "0.000000"),
                ("BT.2020 Area Ratio (%)",   "bt2020_area_ratio", "0.00"),
                ("Measured Area",            "meas_area",         "0.000000"),
            ]
            for label, key, fmt in stat_items:
                value = gamut_stats.get(key)
                lc = ws.cell(row=stat_row, column=1, value=label)
                lc.font = _BOLD_FONT
                lc.alignment = _CENTER
                lc.border = _BORDER
                lc.fill = _stat_fill
                vc = ws.cell(row=stat_row, column=2,
                             value=round(value, 6) if value is not None else None)
                vc.number_format = fmt
                vc.alignment = _CENTER
                vc.border = _BORDER
                vc.fill = _stat_fill
                stat_row += 1

        # u'v' gamut 차트
        if results.get("red") and results.get("green") and results.get("blue"):
            self._add_gamut_chart(ws, results, f"A{stat_row + 2}")

        path = file_path or _default_path(brand, model, "Gamut")
        wb.save(path)
        return path

    # ── 4. Contrast Ratio ─────────────────────────────────────────────────────

    def export_contrast(
        self,
        results: Dict[float, MeasureResult],
        brand: str,
        model: str,
        file_path: str | None = None,
    ) -> str:
        """윈도우 크기별 Lv + 명암비(CR) 자동 계산."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _add_info_sheet(wb, brand, model, "Contrast Ratio")

        ws = wb.create_sheet("ContrastRatio")

        meas_headers = [h for h, *_ in _MEAS_COLS]
        meas_widths  = [w for *_, w in _MEAS_COLS]

        # 0% window = solid white (peak Lv) / 100% window = full black window (min Lv)
        # CR = Lv_white(0% win) / Lv_black(100% win)
        ref_lv = results.get(0.0, None)
        ref_lv_val = ref_lv.Lv if ref_lv else None

        headers = ["Window (%)", "Lv (cd/m²)", "CR (White/Lv)"] + meas_headers
        _write_header_row(ws, 1, headers)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        _set_col_widths(ws, meas_widths, col_offset=3)
        _freeze(ws, "B2")

        sorted_items = sorted(results.items(), reverse=True)  # Full White first, then 100%→14.1%

        for ri, (win_size, r) in enumerate(sorted_items, 2):
            if ref_lv_val and r.Lv > 0 and win_size > 0.0:
                cr_val = round(ref_lv_val / r.Lv, 1)
            else:
                cr_val = None

            row_fill = _GREEN_FILL if win_size == 0.0 else None

            win_label = "Full White" if win_size == 0.0 else win_size
            win_fmt = "@" if win_size == 0.0 else _FMT_RATIO
            for ci, (val, fmt) in enumerate(
                [(win_label, win_fmt), (r.Lv, _FMT_LV), (cr_val, _FMT_RATIO)], 1
            ):
                c = ws.cell(row=ri, column=ci, value=val)
                c.number_format = fmt
                c.alignment = _CENTER
                c.border = _BORDER
                # CR 열(3)은 값 기반 색상, 나머지는 행 색상 또는 기본
                if ci == 3 and cr_val is not None:
                    c.fill = _cr_fill(cr_val)
                    c.font = _cr_font(cr_val)
                elif row_fill:
                    c.fill = row_fill
                    c.font = _BASE_FONT
                else:
                    c.font = _BASE_FONT

            _write_meas_row(ws, ri, r, col_offset=3)
            if row_fill:
                for ci2 in range(4, len(_MEAS_COLS) + 4):
                    ws.cell(row=ri, column=ci2).fill = row_fill

        path = file_path or _default_path(brand, model, "ContrastRatio")
        wb.save(path)
        return path

    # ── 5. All-session combined export ───────────────────────────────────────

    def export_all_session(
        self,
        brand: str,
        model: str,
        session_swing:    Dict[str, Any],   # "SDR_Vivid" → [MeasureResult]
        session_loading:  Dict[str, Any],   # "SDR_Vivid" → {apl → [MeasureResult]}
        session_gamut:    Dict[str, Any],   # "SDR"/"HDR" → {color → MeasureResult}
        session_contrast: Dict[str, Any],   # "SDR"/"HDR" → {side → MeasureResult}
        file_path: str = "",
        serial_number: str = "",
        sw_version: str = "",
        sw_codename: str = "",
    ) -> str:
        """{brand}_{model}_all.xlsx — 모든 측정 결과를 탭으로 나눠 통합 저장."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _add_info_sheet(wb, brand, model, "All Sessions",
                        serial_number=serial_number,
                        sw_version=sw_version,
                        sw_codename=sw_codename)

        meas_headers = [h for h, *_ in _MEAS_COLS]
        meas_widths  = [w for *_, w in _MEAS_COLS]

        # ── 휘도 스윙 (피벗: 행=#, 열=모드, 값=Lv) ──────────────────────────
        if session_swing:
            swing_keys = sorted(k for k, v in session_swing.items() if v)
            if swing_keys:
                max_n = max(len(session_swing[k]) for k in swing_keys)
                ws_sw = wb.create_sheet("LumSwing")
                _write_header_row(ws_sw, 1, ["#"] + swing_keys)
                ws_sw.column_dimensions["A"].width = 6
                for ci in range(2, len(swing_keys) + 2):
                    ws_sw.column_dimensions[get_column_letter(ci)].width = 14
                _freeze(ws_sw, "B2")
                for i in range(max_n):
                    ri = i + 2
                    idx_c = ws_sw.cell(row=ri, column=1, value=i + 1)
                    idx_c.alignment = _CENTER
                    idx_c.border = _BORDER
                    idx_c.font = _BASE_FONT
                    for ci, key in enumerate(swing_keys, 2):
                        row_list = session_swing.get(key, [])
                        val = row_list[i].Lv if i < len(row_list) else None
                        c = ws_sw.cell(row=ri, column=ci, value=val)
                        c.number_format = _FMT_LV
                        c.alignment = _CENTER
                        c.border = _BORDER
                        c.font = _BASE_FONT
                # 통계
                stat_row = max_n + 3
                ws_sw.cell(row=stat_row, column=1, value="Statistics").font = _BOLD_FONT
                stat_row += 1
                for label in ["Count", "Mean Lv", "Max Lv", "Min Lv"]:
                    ws_sw.cell(row=stat_row, column=1, value=label).font = _BOLD_FONT
                    for ci, key in enumerate(swing_keys, 2):
                        lvs = [r.Lv for r in session_swing.get(key, [])]
                        if not lvs:
                            continue
                        if   label == "Count":   val = len(lvs)
                        elif label == "Mean Lv": val = round(statistics.mean(lvs), 3)
                        elif label == "Max Lv":  val = round(max(lvs), 3)
                        else:                    val = round(min(lvs), 3)
                        c = ws_sw.cell(row=stat_row, column=ci, value=val)
                        c.number_format = _FMT_LV
                        c.alignment = _CENTER
                        c.border = _BORDER
                    stat_row += 1
                if max_n >= 2:
                    self._add_swing_pivot_chart(ws_sw, swing_keys, max_n)

        # ── APL 로딩 Summary ──────────────────────────────────────────────────
        if session_loading:
            cases    = sorted(session_loading.keys())
            all_apls = sorted({apl for cd in session_loading.values() for apl in cd})
            ws_sum = wb.create_sheet("Loading_Summary")
            hdr = ["APL (%)"]
            for c in cases:
                hdr += [f"{c} Avg", f"{c} Max", f"{c} Min"]
            _write_header_row(ws_sum, 1, hdr)
            ws_sum.column_dimensions["A"].width = 10
            for ci in range(2, len(hdr) + 1):
                ws_sum.column_dimensions[get_column_letter(ci)].width = 14
            for ri, apl in enumerate(all_apls, 2):
                ws_sum.cell(row=ri, column=1, value=apl).alignment = _CENTER
                ci = 2
                for case in cases:
                    lvs = [r.Lv for r in session_loading.get(case, {}).get(apl, [])]
                    avg = round(statistics.mean(lvs), 3) if lvs else None
                    mx  = round(max(lvs), 3)             if lvs else None
                    mn  = round(min(lvs), 3)             if lvs else None
                    for val in [avg, mx, mn]:
                        c2 = ws_sum.cell(row=ri, column=ci, value=val)
                        c2.number_format = _FMT_LV
                        c2.alignment = _CENTER
                        c2.border = _BORDER
                        ci += 1
            _freeze(ws_sum, "B2")
            # APL 로딩 차트 (3열 형식: Avg/Max/Min)
            if all_apls:
                self._add_apl_chart(ws_sum, all_apls, cases, len(all_apls),
                                    cols_per_case=3)
            # 케이스별 Raw 시트
            for case, apl_dict in sorted(session_loading.items()):
                ws_raw = wb.create_sheet(f"Loading_{case}"[:31])
                headers_raw = ["APL (%)", "#"] + meas_headers
                _write_header_row(ws_raw, 1, headers_raw)
                _set_col_widths(ws_raw, [10, 6] + meas_widths)
                _freeze(ws_raw, "C2")
                ri = 2
                for apl in sorted(apl_dict):
                    for idx, r in enumerate(apl_dict[apl], 1):
                        ws_raw.cell(row=ri, column=1, value=apl).alignment = _CENTER
                        ws_raw.cell(row=ri, column=2, value=idx).alignment = _CENTER
                        _write_meas_row(ws_raw, ri, r, col_offset=2)
                        ri += 1

        # ── 색재현율 ──────────────────────────────────────────────────────────
        color_order = ["red", "green", "blue", "white", "black"]
        color_fills_gamut = {
            "red":   PatternFill("solid", fgColor="FCE4D6"),
            "green": PatternFill("solid", fgColor="E2EFDA"),
            "blue":  PatternFill("solid", fgColor="DAE8FC"),
            "white": PatternFill("solid", fgColor="F2F2F2"),
            "black": PatternFill("solid", fgColor="D9D9D9"),
        }
        for mode_key, gamut_results in sorted(session_gamut.items()):
            if not gamut_results:
                continue
            ws = wb.create_sheet(f"Gamut_{mode_key}"[:31])
            _write_header_row(ws, 1, ["Color"] + meas_headers)
            _set_col_widths(ws, [8] + meas_widths)
            _freeze(ws, "B2")
            for ri, color in enumerate(color_order, 2):
                r = gamut_results.get(color)
                fill = color_fills_gamut.get(color)
                lc = ws.cell(row=ri, column=1, value=color.capitalize())
                lc.font = _BOLD_FONT
                lc.alignment = _CENTER
                lc.border = _BORDER
                if fill:
                    lc.fill = fill
                if r:
                    _write_meas_row(ws, ri, r, col_offset=1)
                    if fill:
                        for ci in range(2, len(meas_headers) + 2):
                            ws.cell(row=ri, column=ci).fill = fill
            # gamut 통계 + 차트
            stat_row_g = len(color_order) + 3
            r_r = gamut_results.get("red");  r_g = gamut_results.get("green")
            r_b = gamut_results.get("blue"); w_r = gamut_results.get("white")
            bk_r = gamut_results.get("black")
            if w_r and bk_r and bk_r.Lv > 0:
                ws.cell(row=stat_row_g, column=1, value="Contrast Ratio").font = _BOLD_FONT
                c2 = ws.cell(row=stat_row_g, column=2, value=round(w_r.Lv / bk_r.Lv, 1))
                c2.number_format = _FMT_RATIO
                c2.font = Font(bold=True, color="C55A11", size=11)
                stat_row_g += 1
            if r_r and r_g and r_b:
                from .gamut_utils import calc_gamut_stats as _cgs
                _stats = _cgs((r_r.u_prime, r_r.v_prime),
                               (r_g.u_prime, r_g.v_prime),
                               (r_b.u_prime, r_b.v_prime))
                _sf = PatternFill("solid", fgColor="EBF3FB")
                for _lbl, _val, _fmt in [
                    ("DCI-P3 Overlap (%)",  _stats["dci_overlap"],   "0.00"),
                    ("DCI-P3 Area",         _stats["dci_area"],      "0.000000"),
                    ("BT.2020 Overlap (%)", _stats["bt2020_overlap"],"0.00"),
                    ("BT.2020 Area",        _stats["bt2020_area"],   "0.000000"),
                    ("Measured Area",       _stats["meas_area"],     "0.000000"),
                ]:
                    lc2 = ws.cell(row=stat_row_g, column=1, value=_lbl)
                    lc2.font = _BOLD_FONT; lc2.alignment = _CENTER
                    lc2.border = _BORDER; lc2.fill = _sf
                    vc2 = ws.cell(row=stat_row_g, column=2, value=round(_val, 6))
                    vc2.number_format = _fmt; vc2.alignment = _CENTER
                    vc2.border = _BORDER; vc2.fill = _sf
                    stat_row_g += 1
                self._add_gamut_chart(ws, gamut_results, f"A{stat_row_g + 2}")

        # ── 명암비 ────────────────────────────────────────────────────────────
        for mode_key, contrast_results in sorted(session_contrast.items()):
            if not contrast_results:
                continue
            ws = wb.create_sheet(f"Contrast_{mode_key}"[:31])
            _write_header_row(ws, 1, ["Black H/V (%)", "Lv (cd/m²)", "CR (White/Lv)"] + meas_headers)
            _set_col_widths(ws, [14, 14, 14] + meas_widths)
            _freeze(ws, "B2")
            ref_lv = contrast_results[0.0].Lv if 0.0 in contrast_results else None
            for ri, side in enumerate(sorted(contrast_results, reverse=True), 2):
                r = contrast_results[side]
                cr = round(ref_lv / r.Lv, 1) if (ref_lv and r.Lv > 0 and side > 0.0) else None
                side_label = "Full White" if side == 0.0 else side
                side_fmt   = "@" if side == 0.0 else _FMT_RATIO
                row_fill   = _GREEN_FILL if side == 0.0 else None
                for ci, (val, fmt) in enumerate(
                    [(side_label, side_fmt), (r.Lv, _FMT_LV), (cr, _FMT_RATIO)], 1
                ):
                    c2 = ws.cell(row=ri, column=ci, value=val)
                    c2.number_format = fmt
                    c2.alignment = _CENTER
                    c2.border = _BORDER
                    if ci == 3 and cr is not None:
                        c2.fill = _cr_fill(cr)
                        c2.font = _cr_font(cr)
                    elif row_fill:
                        c2.fill = row_fill
                        c2.font = _BASE_FONT
                    else:
                        c2.font = _BASE_FONT
                _write_meas_row(ws, ri, r, col_offset=3)
                if row_fill:
                    for ci2 in range(4, len(meas_headers) + 4):
                        ws.cell(row=ri, column=ci2).fill = row_fill

        # ── 요약 시트 추가 ────────────────────────────────────────────────────
        entry = self._build_summary_entry(session_loading, session_gamut, session_contrast)
        self._add_competitor_sheet(wb, brand, model, entry)
        self._add_optical_sheet(wb, brand, model, entry)

        path = file_path or _default_path(brand, model, "all")
        wb.save(path)
        return path

    # ── 요약 데이터 계산 ──────────────────────────────────────────────────────

    def _build_summary_entry(
        self,
        session_loading:  Dict[str, Any],
        session_gamut:    Dict[str, Any],
        session_contrast: Dict[str, Any],
    ) -> Dict[str, Any]:
        """session 데이터에서 경쟁사/광학 표에 필요한 지표를 추출해 반환."""
        entry: Dict[str, Any] = {
            "hdr_10": None, "hdr_100": None,
            "sdr_10": None, "sdr_100": None,
            "sdr_vivid_10":    None, "sdr_vivid_100":    None,
            "sdr_standard_10": None, "sdr_standard_100": None,
            "hdr_vivid_10":    None, "hdr_vivid_100":    None,
            "hdr_standard_10": None, "hdr_standard_100": None,
            "hdr_cinema_10":   None, "hdr_cinema_100":   None,
            "contrast_ratio": None, "black_lv": None,
            "dci_overlap": None, "bt2020_overlap": None,
        }

        # ── 로딩: 각 케이스별 APL 10%/100% 최대 Lv ───────────────────────
        for case_key, apl_dict in session_loading.items():
            upper = case_key.upper()
            is_hdr    = "HDR" in upper
            is_cinema = "CINEMA" in upper
            is_std    = "STANDARD" in upper or "_STD_" in upper

            lv_10  = self._agg_lv(apl_dict.get(10,  []))
            lv_100 = self._agg_lv(apl_dict.get(100, []))

            if is_hdr:
                if lv_10  is not None: entry["hdr_10"]  = max(entry["hdr_10"]  or 0, lv_10)
                if lv_100 is not None: entry["hdr_100"] = max(entry["hdr_100"] or 0, lv_100)
                pfx = "hdr_cinema" if is_cinema else ("hdr_standard" if is_std else "hdr_vivid")
            else:
                if lv_10  is not None: entry["sdr_10"]  = max(entry["sdr_10"]  or 0, lv_10)
                if lv_100 is not None: entry["sdr_100"] = max(entry["sdr_100"] or 0, lv_100)
                pfx = "sdr_standard" if is_std else "sdr_vivid"

            if lv_10  is not None: entry[f"{pfx}_10"]  = lv_10
            if lv_100 is not None: entry[f"{pfx}_100"] = lv_100

        # ── 색재현율: SDR/HDR 중 첫 번째 gamut 결과 사용 ──────────────────
        for _mode, gamut_results in session_gamut.items():
            r_r = gamut_results.get("red")
            r_g = gamut_results.get("green")
            r_b = gamut_results.get("blue")
            if r_r and r_g and r_b:
                stats = calc_gamut_stats(
                    (r_r.u_prime, r_r.v_prime),
                    (r_g.u_prime, r_g.v_prime),
                    (r_b.u_prime, r_b.v_prime),
                )
                entry["dci_overlap"]   = round(stats.get("dci_overlap",   0.0), 1)
                entry["bt2020_overlap"] = round(stats.get("bt2020_overlap", 0.0), 1)
                break

        # ── 명암비: SDR/HDR 중 첫 번째 contrast 결과 사용 ────────────────
        # CR = max(CR at 50%, 20%, 14.1%) — best contrast among window sizes
        for _mode, contrast_results in session_contrast.items():
            white_r = contrast_results.get(0.0)
            if white_r:
                best_cr: float | None = None
                best_black_r = None
                for side_pct in [50.0, 20.0, 14.1]:
                    r = contrast_results.get(side_pct)
                    if r and r.Lv > 0:
                        cr = white_r.Lv / r.Lv
                        if best_cr is None or cr > best_cr:
                            best_cr = cr
                            best_black_r = r
                if best_cr is not None:
                    entry["contrast_ratio"] = round(best_cr, 0)
                if best_black_r is not None:
                    entry["black_lv"] = round(best_black_r.Lv, 4)
            break

        return entry

    @staticmethod
    def _agg_lv(results: List[Any]) -> float | None:
        """측정 결과 리스트에서 최대 Lv 반환 (최대값 기준)."""
        lvs = [r.Lv for r in results if r is not None]
        return round(max(lvs), 1) if lvs else None

    def _add_competitor_sheet(
        self, wb: Any, brand: str, model: str, entry: Dict[str, Any]
    ) -> None:
        """경쟁사 비교 장표 시트 추가."""
        ws = wb.create_sheet("경쟁사비교")
        label_col = f"{brand}_{model}"

        ROW_LABELS = [
            ("White 휘도[nit]", "HDR 10%"),
            ("White 휘도[nit]", "HDR 100%"),
            ("White 휘도[nit]", "SDR 10%"),
            ("White 휘도[nit]", "SDR 100%"),
            ("White 휘도[nit]", "Contrast Ratio"),
            ("White 휘도[nit]", "Black"),
            ("Color Gamut[%]",  "DCI-P3 (%)"),
            ("Color Gamut[%]",  "BT.2020 (%)"),
        ]
        KEYS = [
            "hdr_10", "hdr_100", "sdr_10", "sdr_100",
            "contrast_ratio", "black_lv", "dci_overlap", "bt2020_overlap",
        ]
        FMTS = [
            "0", "0", "0", "0",
            "0", "0.0000", "0.0", "0.0",
        ]

        _write_header_row(ws, 1, ["구분", "항목", label_col])
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 16

        prev_section = None
        for ri, ((section, item), key, fmt) in enumerate(zip(ROW_LABELS, KEYS, FMTS), 2):
            val = entry.get(key)
            c_sec = ws.cell(row=ri, column=1, value=section if section != prev_section else "")
            c_sec.alignment = _CENTER; c_sec.border = _BORDER; c_sec.font = _BOLD_FONT
            prev_section = section
            c_itm = ws.cell(row=ri, column=2, value=item)
            c_itm.alignment = _CENTER; c_itm.border = _BORDER; c_itm.font = _BASE_FONT
            c_val = ws.cell(row=ri, column=3, value=val)
            c_val.alignment = _CENTER; c_val.border = _BORDER; c_val.font = _BASE_FONT
            if val is not None:
                c_val.number_format = fmt

    def _add_optical_sheet(
        self, wb: Any, brand: str, model: str, entry: Dict[str, Any]
    ) -> None:
        """광학 측정 데이터 시트 추가."""
        ws = wb.create_sheet("광학측정")
        label_col = f"{brand}_{model}"

        OPTICAL_LABELS = [
            ("휘도",        "Vivid SDR 10% / 100%"),
            ("휘도",        "Standard SDR 10% / 100%"),
            ("휘도",        "Vivid HDR 10% / 100%"),
            ("휘도",        "Standard HDR 10% / 100%"),
            ("휘도",        "Cinema HDR 10% / 100%"),
            ("Contrast",    "Black (Ratio)"),
            ("Color Gamut", "DCI-P3 (%)"),
            ("Color Gamut", "BT.2020 (%)"),
        ]
        OPTICAL_KEYS: List[tuple] = [
            ("sdr_vivid_10",    "sdr_vivid_100"),
            ("sdr_standard_10", "sdr_standard_100"),
            ("hdr_vivid_10",    "hdr_vivid_100"),
            ("hdr_standard_10", "hdr_standard_100"),
            ("hdr_cinema_10",   "hdr_cinema_100"),
            ("contrast_ratio",  None),
            ("dci_overlap",     None),
            ("bt2020_overlap",  None),
        ]
        FMTS = ["0", "0", "0", "0", "0", "0", "0.0", "0.0"]

        _write_header_row(ws, 1, ["구분", "항목", label_col])
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 20

        prev_section = None
        for ri, ((section, item), (k10, k100), fmt) in enumerate(
            zip(OPTICAL_LABELS, OPTICAL_KEYS, FMTS), 2
        ):
            v10  = entry.get(k10)
            v100 = entry.get(k100) if k100 else None

            if v10 is not None and v100 is not None:
                cell_val = f"{int(v10) if fmt=='0' else v10} / {int(v100) if fmt=='0' else v100}"
            elif v10 is not None:
                cell_val: Any = int(v10) if fmt == "0" else v10
            else:
                cell_val = None

            c_sec = ws.cell(row=ri, column=1, value=section if section != prev_section else "")
            c_sec.alignment = _CENTER; c_sec.border = _BORDER; c_sec.font = _BOLD_FONT
            prev_section = section
            c_itm = ws.cell(row=ri, column=2, value=item)
            c_itm.alignment = _CENTER; c_itm.border = _BORDER; c_itm.font = _BASE_FONT
            c_val = ws.cell(row=ri, column=3, value=cell_val)
            c_val.alignment = _CENTER; c_val.border = _BORDER; c_val.font = _BASE_FONT

    # ── 6. Report Template ────────────────────────────────────────────────────

    def export_report_template(
        self,
        brand: str,
        model: str,
        gamut_results: Dict[str, "MeasureResult"] | None = None,
        lum_loading_results: Dict[str, Dict[int, List["MeasureResult"]]] | None = None,
        contrast_results: Dict[float, "MeasureResult"] | None = None,
        file_path: str | None = None,
    ) -> str:
        """모든 측정 결과를 하나의 통합 리포트 워크북으로 저장.

        포함 시트:
          - Info      : 기본 정보
          - Summary   : 측정 항목별 핵심 지표 요약
          - Gamut     : 색재현율 측정값 + 통계
          - LumLoading: APL vs Lv 요약 + 차트
          - Contrast  : 윈도우별 명암비
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _add_info_sheet(wb, brand, model, "Report")

        # ── Summary 시트 ──────────────────────────────────────────────────────
        ws_sum = wb.create_sheet("Summary")
        ws_sum.column_dimensions["A"].width = 32
        ws_sum.column_dimensions["B"].width = 20
        ws_sum.column_dimensions["C"].width = 16

        _write_header_row(ws_sum, 1, ["항목", "측정값", "비고"])

        rows_data: list[tuple[str, object, str]] = []

        # Gamut 통계
        if gamut_results:
            r_r = gamut_results.get("red")
            r_g = gamut_results.get("green")
            r_b = gamut_results.get("blue")
            if r_r and r_g and r_b:
                stats = calc_gamut_stats(
                    (r_r.u_prime, r_r.v_prime),
                    (r_g.u_prime, r_g.v_prime),
                    (r_b.u_prime, r_b.v_prime),
                )
                rows_data.append(("DCI-P3 Overlap (%)",        stats["dci_overlap"],        ""))
                rows_data.append(("DCI-P3 Intersection Area",  stats["dci_inter_area"],     "u'v' area"))
                rows_data.append(("DCI-P3 Area",               stats["dci_area"],           "u'v' area"))
                rows_data.append(("DCI-P3 Area Ratio (%)",     stats["dci_area_ratio"],     "meas/DCI-P3"))
                rows_data.append(("BT.2020 Overlap (%)",       stats["bt2020_overlap"],     ""))
                rows_data.append(("BT.2020 Intersection Area", stats["bt2020_inter_area"],  "u'v' area"))
                rows_data.append(("BT.2020 Area",              stats["bt2020_area"],        "u'v' area"))
                rows_data.append(("BT.2020 Area Ratio (%)",    stats["bt2020_area_ratio"],  "meas/BT.2020"))
                rows_data.append(("Measured Area",             stats["meas_area"],          "u'v' area"))
            if gamut_results.get("white") and gamut_results.get("black"):
                w_lv  = gamut_results["white"].Lv
                bk_lv = gamut_results["black"].Lv
                if bk_lv > 0:
                    rows_data.append(("Gamut CR (White/Black)", round(w_lv / bk_lv, 1), ""))

        # LumLoading 통계 — 케이스별 Peak Lv (APL 10%)
        if lum_loading_results:
            for case, case_data in sorted(lum_loading_results.items()):
                apl10 = case_data.get(10, [])
                if apl10:
                    peak = max(r.Lv for r in apl10)
                    rows_data.append((f"LumLoading Case {case} Peak Lv (APL10)", round(peak, 3), "cd/m²"))

        # Contrast — CR = max(CR at 50%, 20%, 14.1%)
        if contrast_results:
            ref = contrast_results.get(0.0)
            if ref:
                best_cr: float | None = None
                for side_pct in [50.0, 20.0, 14.1]:
                    r = contrast_results.get(side_pct)
                    if r and r.Lv > 0:
                        cr = ref.Lv / r.Lv
                        if best_cr is None or cr > best_cr:
                            best_cr = cr
                if best_cr is not None:
                    rows_data.append(("Contrast Ratio (max 50%/20%/14.1%)", round(best_cr, 1), ""))

        for ri, (label, val, note) in enumerate(rows_data, 2):
            fill = _GREEN_FILL if ri % 2 == 0 else None
            for ci, v in enumerate([label, val, note], 1):
                c = ws_sum.cell(row=ri, column=ci, value=v)
                c.alignment = _CENTER
                c.border = _BORDER
                c.font = _BASE_FONT
                if fill:
                    c.fill = fill

        # ── Gamut 시트 ────────────────────────────────────────────────────────
        if gamut_results:
            ws_g = wb.create_sheet("Gamut")
            color_order = ["red", "green", "blue", "white", "black"]
            color_fills = {
                "red":   PatternFill("solid", fgColor="FCE4D6"),
                "green": PatternFill("solid", fgColor="E2EFDA"),
                "blue":  PatternFill("solid", fgColor="DAE8FC"),
                "white": PatternFill("solid", fgColor="F2F2F2"),
                "black": PatternFill("solid", fgColor="D9D9D9"),
            }
            headers = ["Color"] + [h for h, *_ in _MEAS_COLS]
            widths  = [8] + [w for *_, w in _MEAS_COLS]
            _write_header_row(ws_g, 1, headers)
            _set_col_widths(ws_g, widths)
            ws_g.column_dimensions["A"].width = 28   # 통계 레이블 여유 확보
            _freeze(ws_g, "B2")
            for ri, color in enumerate(color_order, 2):
                r = gamut_results.get(color)
                fill = color_fills.get(color)
                lc = ws_g.cell(row=ri, column=1, value=color.capitalize())
                lc.font = _BOLD_FONT
                lc.alignment = _CENTER
                lc.border = _BORDER
                if fill:
                    lc.fill = fill
                if r:
                    _write_meas_row(ws_g, ri, r, col_offset=1)
                    if fill:
                        for ci in range(2, len(_MEAS_COLS) + 2):
                            ws_g.cell(row=ri, column=ci).fill = fill

            # Gamut 통계 블록
            r_r = gamut_results.get("red")
            r_g = gamut_results.get("green")
            r_b = gamut_results.get("blue")
            if r_r and r_g and r_b:
                stats = calc_gamut_stats(
                    (r_r.u_prime, r_r.v_prime),
                    (r_g.u_prime, r_g.v_prime),
                    (r_b.u_prime, r_b.v_prime),
                )
                stat_row = len(color_order) + 3
                _stat_fill = PatternFill("solid", fgColor="EBF3FB")
                for label, key, fmt in [
                    ("DCI-P3 Overlap (%)",        "dci_overlap",       "0.00"),
                    ("DCI-P3 Intersection Area",  "dci_inter_area",    "0.000000"),
                    ("DCI-P3 Area",               "dci_area",          "0.000000"),
                    ("DCI-P3 Area Ratio (%)",     "dci_area_ratio",    "0.00"),
                    ("BT.2020 Overlap (%)",       "bt2020_overlap",    "0.00"),
                    ("BT.2020 Intersection Area", "bt2020_inter_area", "0.000000"),
                    ("BT.2020 Area",              "bt2020_area",       "0.000000"),
                    ("BT.2020 Area Ratio (%)",    "bt2020_area_ratio", "0.00"),
                    ("Measured Area",             "meas_area",         "0.000000"),
                ]:
                    c_label = ws_g.cell(row=stat_row, column=1, value=label)
                    c_label.font = _BOLD_FONT
                    c_label.alignment = _CENTER
                    c_label.border = _BORDER
                    c_label.fill = _stat_fill
                    c_val = ws_g.cell(row=stat_row, column=2, value=round(stats[key], 6))
                    c_val.number_format = fmt
                    c_val.alignment = _CENTER
                    c_val.border = _BORDER
                    c_val.fill = _stat_fill
                    stat_row += 1

        # ── LumLoading 시트 ───────────────────────────────────────────────────
        if lum_loading_results:
            import statistics as _stats
            ws_ll = wb.create_sheet("LumLoading")
            cases    = sorted(lum_loading_results.keys())
            all_apls = sorted({apl for cd in lum_loading_results.values() for apl in cd})
            hdr = ["APL (%)"]
            for case in cases:
                hdr += [f"Case {case} Avg Lv", f"Case {case} Max Lv",
                        f"Case {case} Min Lv", f"Case {case} StdDev"]
            _write_header_row(ws_ll, 1, hdr)
            ws_ll.column_dimensions["A"].width = 10
            for ci in range(2, len(hdr) + 1):
                ws_ll.column_dimensions[get_column_letter(ci)].width = 14
            for ri, apl in enumerate(all_apls, 2):
                ws_ll.cell(row=ri, column=1, value=apl).alignment = _CENTER
                ci = 2
                for case in cases:
                    lvs = [r.Lv for r in lum_loading_results.get(case, {}).get(apl, [])]
                    avg = round(_stats.mean(lvs), 3)   if lvs else None
                    mx  = round(max(lvs), 3)            if lvs else None
                    mn  = round(min(lvs), 3)            if lvs else None
                    std = round(_stats.stdev(lvs), 4)   if len(lvs) > 1 else (0 if lvs else None)
                    for val, fmt in [(avg, _FMT_LV), (mx, _FMT_LV), (mn, _FMT_LV), (std, _FMT_LV)]:
                        c = ws_ll.cell(row=ri, column=ci, value=val)
                        c.number_format = fmt
                        c.alignment = _CENTER
                        c.border = _BORDER
                        ci += 1
            _freeze(ws_ll, "B2")
            self._add_apl_chart(ws_ll, all_apls, cases, len(all_apls), brand=brand)

        # ── Contrast 시트 ─────────────────────────────────────────────────────
        if contrast_results:
            ws_cr = wb.create_sheet("Contrast")
            ref_lv_val = contrast_results.get(0.0)
            ref_lv_val = ref_lv_val.Lv if ref_lv_val else None
            meas_headers = [h for h, *_ in _MEAS_COLS]
            meas_widths  = [w for *_, w in _MEAS_COLS]
            headers = ["Window (%)", "Lv (cd/m²)", "CR (White/Lv)"] + meas_headers
            _write_header_row(ws_cr, 1, headers)
            ws_cr.column_dimensions["A"].width = 12
            ws_cr.column_dimensions["B"].width = 14
            ws_cr.column_dimensions["C"].width = 16
            _set_col_widths(ws_cr, meas_widths, col_offset=3)
            _freeze(ws_cr, "B2")
            for ri, (win_size, r) in enumerate(sorted(contrast_results.items(), reverse=True), 2):
                cr_val = round(ref_lv_val / r.Lv, 1) if (ref_lv_val and r.Lv > 0 and win_size > 0.0) else None
                fill = _GREEN_FILL if win_size == 0.0 else (
                       _YELLOW_FILL if win_size <= 20.0 else None)
                win_label = "Full White" if win_size == 0.0 else win_size
                win_fmt = "@" if win_size == 0.0 else _FMT_RATIO
                for ci, (val, fmt) in enumerate(
                    [(win_label, win_fmt), (r.Lv, _FMT_LV), (cr_val, _FMT_RATIO)], 1
                ):
                    c = ws_cr.cell(row=ri, column=ci, value=val)
                    c.number_format = fmt
                    c.alignment = _CENTER
                    c.border = _BORDER
                    c.font = _BOLD_FONT if ci == 3 else _BASE_FONT
                    if fill:
                        c.fill = fill
                _write_meas_row(ws_cr, ri, r, col_offset=3)
                if fill:
                    for ci2 in range(4, len(_MEAS_COLS) + 4):
                        ws_cr.cell(row=ri, column=ci2).fill = fill

        path = file_path or _default_path(brand, model, "Report")
        wb.save(path)
        return path
